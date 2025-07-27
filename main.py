import os
import sys
import io
import requests
import pandas as pd
import matplotlib.pyplot as plt
import smtplib
import msal
import base64
import traceback
import time
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Import openpyxl for advanced Excel formatting
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Global processor variable
global_processor = None

# SharePoint Configuration
SHAREPOINT_CONFIG = {
    'tenant_id': '81060475-7e7f-4ede-8d8d-bf61f53ca528',
    'client_id': '076541aa-c734-405e-8518-ed52b67f8cbd',
    'authority': 'https://login.microsoftonline.com/81060475-7e7f-4ede-8d8d-bf61f53ca528',
    'scopes': ['https://graph.microsoft.com/Sites.ReadWrite.All'],
    'site_name': 'MCH.MMB.QA',
    'base_url': 'masangroup.sharepoint.com'
}

# SharePoint File ID for "CIP plan.xlsx"
CIP_PLAN_FILE_ID = '8C90FB38-DA8C-59CC-547D-53BEA1C8B16D'

# Excel formatting styles
def create_excel_styles():
    """Create professional Excel styles for different data types"""
    
    # Header style
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_style.border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Normal data style
    normal_style = NamedStyle(name="normal_style")
    normal_style.font = Font(name='Arial', size=10)
    normal_style.border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    normal_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Date style
    date_style = NamedStyle(name="date_style")
    date_style.font = Font(name='Arial', size=10)
    date_style.border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    date_style.alignment = Alignment(horizontal='center', vertical='center')
    date_style.number_format = 'DD/MM/YYYY'
    
    # Center align style
    center_style = NamedStyle(name="center_style")
    center_style.font = Font(name='Arial', size=10)
    center_style.border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    center_style.alignment = Alignment(horizontal='center', vertical='center')
    
    # Status styles with different colors
    status_styles = {
        'Bình thường': {
            'fill': PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid"),
            'font': Font(name='Arial', size=10, color="2D5016", bold=True)
        },
        'Sắp đến hạn': {
            'fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'font': Font(name='Arial', size=10, color="7F6000", bold=True)
        },
        'Đến hạn': {
            'fill': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
            'font': Font(name='Arial', size=10, color="9C6500", bold=True)
        },
        'Quá hạn': {
            'fill': PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid"),
            'font': Font(name='Arial', size=10, color="9C0006", bold=True)
        },
        'Chưa có dữ liệu': {
            'fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            'font': Font(name='Arial', size=10, color="7F7F7F")
        },
        'Lỗi': {
            'fill': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            'font': Font(name='Arial', size=10, color="FFFFFF", bold=True)
        }
    }
    
    return header_style, normal_style, date_style, center_style, status_styles

def format_worksheet(worksheet, df, sheet_name, status_styles, center_style):
    """Apply professional formatting to a worksheet"""
    
    # Set column widths based on content and column names
    column_widths = {
        'Khu vực': 18,
        'Thiết bị': 25,
        'Phương pháp': 15,
        'Tần suất (ngày)': 14,
        'Ngày vệ sinh gần nhất': 20,
        'Ngày kế hoạch vệ sinh tiếp theo': 25,
        'Trạng thái': 16,
        'Đang chứa sản phẩm': 20,
        'Người thực hiện': 18,
        'Kết quả': 12,
        'Ghi chú': 35,
        'Ngày vệ sinh': 16
    }
    
    # Auto-adjust column widths
    for idx, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        
        # Set width based on column name or calculate from content
        base_width = column_widths.get(column, 15)
        
        # Calculate max content width for better auto-sizing
        max_length = len(str(column))
        for cell_value in df[column].astype(str):
            if len(cell_value) > max_length:
                max_length = min(len(cell_value), 60)  # Cap at 60 chars
        
        # Use the larger of predefined width or content-based width
        final_width = max(base_width, max_length + 3)
        worksheet.column_dimensions[column_letter].width = final_width
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.style = "header_style"
    
    # Apply data formatting
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            column_name = df.columns[col_idx - 1]
            
            # Apply date formatting to date columns
            if any(date_keyword in column_name.lower() for date_keyword in ['ngày', 'date']):
                cell.style = "date_style"
                
                # Try to parse and format date properly
                if cell.value and str(cell.value).strip() not in ['nan', 'None', '']:
                    try:
                        if isinstance(cell.value, str):
                            # Try to parse the date string
                            date_obj = parse_date(cell.value)
                            if date_obj:
                                cell.value = date_obj
                    except:
                        pass
            
            # Apply status formatting
            elif 'trạng thái' in column_name.lower():
                cell_value = str(cell.value).strip()
                if cell_value in status_styles:
                    cell.fill = status_styles[cell_value]['fill']
                    cell.font = status_styles[cell_value]['font']
                    cell.border = Border(
                        left=Side(style='thin', color='D3D3D3'),
                        right=Side(style='thin', color='D3D3D3'),
                        top=Side(style='thin', color='D3D3D3'),
                        bottom=Side(style='thin', color='D3D3D3')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.style = "normal_style"
            
            # Apply center alignment for specific columns
            elif any(keyword in column_name.lower() for keyword in ['tần suất', 'kết quả', 'frequency']):
                cell.style = center_style
            
            # Apply normal formatting to other cells
            else:
                cell.style = "normal_style"
    
    # Add conditional formatting for critical items (has product + overdue)
    add_critical_formatting(worksheet, df)
    
    # Freeze panes (freeze first row)
    worksheet.freeze_panes = worksheet['A2']
    
    # Add filter to header row
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    
    # Set row heights
    worksheet.row_dimensions[1].height = 35  # Header row
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 25

def add_critical_formatting(worksheet, df):
    """Add conditional formatting for critical items"""
    try:
        # Find status and product columns
        status_col = None
        product_col = None
        
        for idx, col in enumerate(df.columns):
            if 'trạng thái' in col.lower():
                status_col = idx + 1
            elif 'chứa sản phẩm' in col.lower():
                product_col = idx + 1
        
        if status_col and product_col:
            # Rule for equipment with product that is overdue (CRITICAL)
            status_col_letter = get_column_letter(status_col)
            product_col_letter = get_column_letter(product_col)
            
            critical_rule = FormulaRule(
                formula=[f'AND(${status_col_letter}2="Quá hạn", ${product_col_letter}2<>"")'],
                fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                font=Font(color="FFFFFF", bold=True)
            )
            
            worksheet.conditional_formatting.add(
                f'A2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}',
                critical_rule
            )
    except Exception as e:
        print(f"Warning: Could not add critical formatting: {e}")

def add_summary_section(worksheet, df):
    """Add a comprehensive summary section to the Master plan sheet"""
    
    # Find last row with data
    last_row = worksheet.max_row
    
    # Add some spacing
    summary_start_row = last_row + 3
    
    # Add summary header
    summary_cell = worksheet.cell(row=summary_start_row, column=1)
    summary_cell.value = "📊 THỐNG KÊ TỔNG QUAN TÌNH TRẠNG THIẾT BỊ"
    summary_cell.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    summary_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Merge cells for summary header
    worksheet.merge_cells(f'A{summary_start_row}:F{summary_start_row}')
    worksheet.row_dimensions[summary_start_row].height = 35
    
    # Calculate statistics
    if 'Trạng thái' in df.columns:
        status_counts = df['Trạng thái'].value_counts()
        total_equipment = len(df)
        
        # Calculate compliance rate
        compliant = status_counts.get('Bình thường', 0) + status_counts.get('Sắp đến hạn', 0)
        compliance_rate = (compliant / total_equipment * 100) if total_equipment > 0 else 0
        
        # Add main statistics
        stats_start_row = summary_start_row + 2
        
        # Main KPIs
        main_kpis = [
            ['📈 Tỷ lệ tuân thủ (Compliance Rate):', f"{compliance_rate:.1f}%"],
            ['🏭 Tổng số thiết bị:', total_equipment],
            ['🚨 Thiết bị cần chú ý ngay:', status_counts.get('Quá hạn', 0) + status_counts.get('Đến hạn', 0)]
        ]
        
        for i, (label, value) in enumerate(main_kpis):
            row = stats_start_row + i
            
            # Label cell
            label_cell = worksheet.cell(row=row, column=1)
            label_cell.value = label
            label_cell.font = Font(name='Arial', size=12, bold=True)
            
            # Value cell
            value_cell = worksheet.cell(row=row, column=3)
            value_cell.value = value
            value_cell.font = Font(name='Arial', size=12, bold=True, color="C00000")
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Detailed breakdown
        detail_start_row = stats_start_row + len(main_kpis) + 2
        
        # Detail header
        detail_header = worksheet.cell(row=detail_start_row, column=1)
        detail_header.value = "📋 CHI TIẾT THEO TRẠNG THÁI:"
        detail_header.font = Font(name='Arial', size=12, bold=True, color="366092")
        
        stats_data = [
            ['✅ Bình thường:', status_counts.get('Bình thường', 0), 'D4F4DD', '2D5016'],
            ['⚠️ Sắp đến hạn:', status_counts.get('Sắp đến hạn', 0), 'FFF2CC', '7F6000'],
            ['🔶 Đến hạn:', status_counts.get('Đến hạn', 0), 'FFE699', '9C6500'],
            ['🔴 Quá hạn:', status_counts.get('Quá hạn', 0), 'FFCCCB', '9C0006'],
            ['❓ Chưa có dữ liệu:', status_counts.get('Chưa có dữ liệu', 0), 'F2F2F2', '7F7F7F']
        ]
        
        for i, (label, value, bg_color, font_color) in enumerate(stats_data):
            row = detail_start_row + 1 + i
            
            # Label cell
            label_cell = worksheet.cell(row=row, column=1)
            label_cell.value = label
            label_cell.font = Font(name='Arial', size=11, bold=True)
            
            # Value cell
            value_cell = worksheet.cell(row=row, column=3)
            value_cell.value = value
            value_cell.font = Font(name='Arial', size=11, bold=True, color=font_color)
            value_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Percentage cell
            if total_equipment > 0:
                percentage = (value / total_equipment * 100)
                pct_cell = worksheet.cell(row=row, column=4)
                pct_cell.value = f"({percentage:.1f}%)"
                pct_cell.font = Font(name='Arial', size=10, color=font_color)
                pct_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add equipment with product analysis
        if 'Đang chứa sản phẩm' in df.columns:
            product_start_row = detail_start_row + len(stats_data) + 3
            
            # Product analysis header
            product_header = worksheet.cell(row=product_start_row, column=1)
            product_header.value = "🏭 PHÂN TÍCH THIẾT BỊ CHỨA SẢN PHẨM:"
            product_header.font = Font(name='Arial', size=12, bold=True, color="C00000")
            
            # Calculate equipment with product that are overdue
            overdue_with_product = 0
            due_with_product = 0
            
            for idx, row_data in df.iterrows():
                status = str(row_data.get('Trạng thái', '')).strip()
                has_product = str(row_data.get('Đang chứa sản phẩm', '')).strip()
                
                if has_product and has_product not in ['nan', 'None', '']:
                    if status == 'Quá hạn':
                        overdue_with_product += 1
                    elif status == 'Đến hạn':
                        due_with_product += 1
            
            product_stats = [
                ['🚨 Quá hạn + Có sản phẩm (KHẨN CẤP):', overdue_with_product],
                ['⚠️ Đến hạn + Có sản phẩm:', due_with_product]
            ]
            
            for i, (label, value) in enumerate(product_stats):
                row = product_start_row + 1 + i
                
                # Label cell
                label_cell = worksheet.cell(row=row, column=1)
                label_cell.value = label
                label_cell.font = Font(name='Arial', size=11, bold=True)
                
                # Value cell
                value_cell = worksheet.cell(row=row, column=3)
                value_cell.value = value
                if value > 0:
                    value_cell.font = Font(name='Arial', size=11, bold=True, color="C00000")
                    value_cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                else:
                    value_cell.font = Font(name='Arial', size=11, bold=True, color="2D5016")
                    value_cell.fill = PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid")
                
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add timestamp
        timestamp_row = product_start_row + len(product_stats) + 3
        timestamp_cell = worksheet.cell(row=timestamp_row, column=1)
        timestamp_cell.value = f"🕒 Cập nhật lần cuối: {datetime.now().strftime('%d/%m/%Y lúc %H:%M:%S')}"
        timestamp_cell.font = Font(name='Arial', size=10, italic=True, color="7F7F7F")
        
        # Add note
        note_row = timestamp_row + 1
        note_cell = worksheet.cell(row=note_row, column=1)
        note_cell.value = "💡 Lưu ý: Thiết bị có sản phẩm + quá hạn được highlight màu đỏ cần ưu tiên xử lý"
        note_cell.font = Font(name='Arial', size=10, italic=True, color="7F7F7F")

def create_formatted_excel(sheets_data):
    """Create a professionally formatted Excel file"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create and register styles
    header_style, normal_style, date_style, center_style, status_styles = create_excel_styles()
    
    # Add styles to workbook
    wb.add_named_style(header_style)
    wb.add_named_style(normal_style)
    wb.add_named_style(date_style)
    wb.add_named_style(center_style)
    
    # Define sheet order
    sheet_order = ['Master plan', 'Actual result', 'Cleaning History']
    
    # Process sheets in order
    for sheet_name in sheet_order:
        if sheet_name in sheets_data and not sheets_data[sheet_name].empty:
            df = sheets_data[sheet_name]
            
            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply formatting
            format_worksheet(ws, df, sheet_name, status_styles, center_style)
            
            # Add summary information for Master plan sheet
            if sheet_name == 'Master plan' and 'Trạng thái' in df.columns:
                add_summary_section(ws, df)
    
    # Process any remaining sheets not in the order
    for sheet_name, df in sheets_data.items():
        if sheet_name not in sheet_order and not df.empty:
            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply formatting
            format_worksheet(ws, df, sheet_name, status_styles, center_style)
    
    return wb

class GitHubSecretsUpdater:
    """Helper class to update GitHub Secrets using GitHub API"""
    def __init__(self, repo_owner, repo_name, github_token):
        self.repo_owner = repo_owner
        self.repo_name = repo_name
        self.github_token = github_token
        self.api_base = "https://api.github.com"
    
    def get_public_key(self):
        """Get repository public key for encrypting secrets"""
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/actions/secrets/public-key"
        headers = {
            "Authorization": f"token {self.github_token}",
            "Accept": "application/vnd.github.v3+json"
        }
        
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        else:
            raise Exception(f"Failed to get public key: {response.status_code}")
    
    def encrypt_secret(self, public_key, secret_value):
        """Encrypt secret using repository public key"""
        from nacl import encoding, public
        
        public_key_obj = public.PublicKey(public_key.encode("utf-8"), encoding.Base64Encoder())
        sealed_box = public.SealedBox(public_key_obj)
        encrypted = sealed_box.encrypt(secret_value.encode("utf-8"))
        
        return base64.b64encode(encrypted).decode("utf-8")
    
    def update_secret(self, secret_name, secret_value):
        """Update a GitHub secret"""
        try:
            # Get public key
            key_data = self.get_public_key()
            
            # Encrypt secret
            encrypted_value = self.encrypt_secret(key_data["key"], secret_value)
            
            # Update secret
            url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/actions/secrets/{secret_name}"
            headers = {
                "Authorization": f"token {self.github_token}",
                "Accept": "application/vnd.github.v3+json"
            }
            data = {
                "encrypted_value": encrypted_value,
                "key_id": key_data["key_id"]
            }
            
            response = requests.put(url, headers=headers, json=data)
            if response.status_code in [201, 204]:
                print(f"✅ Successfully updated {secret_name}")
                return True
            else:
                print(f"❌ Failed to update {secret_name}: {response.status_code}")
                return False
                
        except Exception as e:
            print(f"❌ Error updating secret: {str(e)}")
            return False

class SharePointCIPProcessor:
    """SharePoint integration for CIP Cleaning automation with improved formatting"""
    
    def __init__(self):
        self.access_token = None
        self.refresh_token = None
        self.base_url = "https://graph.microsoft.com/v1.0"
        self.site_id = None
        self.msal_app = None
        
        # Initialize MSAL app
        self.msal_app = msal.PublicClientApplication(
            SHAREPOINT_CONFIG['client_id'],
            authority=SHAREPOINT_CONFIG['authority']
        )
        
        # Authenticate on initialization
        if not self.authenticate():
            raise Exception("SharePoint authentication failed during initialization")

    def log(self, message):
        """Log with timestamp"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] {message}")
        sys.stdout.flush()

    def authenticate(self):
        """Authenticate using delegation flow with pre-generated tokens"""
        try:
            self.log("🔐 Authenticating with SharePoint...")

            # Get tokens from environment variables
            access_token = os.environ.get('SHAREPOINT_ACCESS_TOKEN')
            refresh_token = os.environ.get('SHAREPOINT_REFRESH_TOKEN')

            if not access_token and not refresh_token:
                self.log("❌ No SharePoint tokens found in environment variables")
                return False

            self.access_token = access_token
            self.refresh_token = refresh_token
            
            if access_token:
                self.log(f"✅ Found access token: {access_token[:30]}...")
                
                # Test token validity
                if self.test_token_validity():
                    self.log("✅ SharePoint access token is valid")
                    return True
                else:
                    self.log("⚠️ SharePoint access token expired, attempting refresh...")
                    
            # Try to refresh token
            if refresh_token:
                if self.refresh_access_token():
                    self.log("✅ SharePoint token refreshed successfully")
                    self.update_github_secrets()
                    return True
                else:
                    self.log("❌ SharePoint token refresh failed")
                    return False
            else:
                self.log("❌ No SharePoint refresh token available")
                return False

        except Exception as e:
            self.log(f"❌ SharePoint authentication error: {str(e)}")
            return False

    def test_token_validity(self):
        """Test if current access token is valid"""
        try:
            headers = self.get_headers()
            response = requests.get(f"{self.base_url}/me", headers=headers, timeout=30)

            if response.status_code == 200:
                user_info = response.json()
                self.log(f"✅ Authenticated to SharePoint as: {user_info.get('displayName', 'Unknown')}")
                return True
            elif response.status_code == 401:
                return False
            else:
                self.log(f"Warning: Unexpected response code: {response.status_code}")
                return False

        except Exception as e:
            self.log(f"Error testing SharePoint token validity: {str(e)}")
            return False

    def refresh_access_token(self):
        """Refresh access token using refresh token with MSAL"""
        try:
            if not self.refresh_token:
                self.log("❌ No refresh token available")
                return False

            self.log("🔄 Attempting to refresh SharePoint token using MSAL...")

            # Use MSAL to refresh token
            result = self.msal_app.acquire_token_by_refresh_token(
                self.refresh_token,
                scopes=SHAREPOINT_CONFIG['scopes']
            )

            if result and "access_token" in result:
                self.access_token = result['access_token']
                if 'refresh_token' in result:
                    self.refresh_token = result['refresh_token']
                    self.log("✅ Got new refresh token")
                
                self.log("✅ SharePoint token refreshed successfully")
                return True
            else:
                error = result.get('error_description', 'Unknown error') if result else 'No result'
                self.log(f"❌ SharePoint token refresh failed: {error}")
                return False

        except Exception as e:
            self.log(f"❌ Error refreshing SharePoint token: {str(e)}")
            return False

    def update_github_secrets(self):
        """Update GitHub Secrets with new tokens"""
        try:
            github_token = os.environ.get('GITHUB_TOKEN')
            if not github_token:
                self.log("⚠️ No GITHUB_TOKEN found, cannot update secrets")
                return False
            
            repo = os.environ.get('GITHUB_REPOSITORY', '')
            if '/' not in repo:
                self.log("⚠️ Invalid GITHUB_REPOSITORY format")
                return False
            
            repo_owner, repo_name = repo.split('/')
            updater = GitHubSecretsUpdater(repo_owner, repo_name, github_token)
            
            # Update access token
            if self.access_token:
                updater.update_secret('SHAREPOINT_ACCESS_TOKEN', self.access_token)
            
            # Update refresh token
            if self.refresh_token:
                updater.update_secret('SHAREPOINT_REFRESH_TOKEN', self.refresh_token)
            
            return True
            
        except Exception as e:
            self.log(f"⚠️ Error updating GitHub Secrets: {str(e)}")
            return False

    def get_headers(self):
        """Get headers for API requests"""
        return {
            'Authorization': f'Bearer {self.access_token}',
            'Content-Type': 'application/json'
        }

    def get_site_id(self):
        """Get SharePoint site ID"""
        try:
            if self.site_id:
                return self.site_id

            url = f"{self.base_url}/sites/{SHAREPOINT_CONFIG['base_url']}:/sites/{SHAREPOINT_CONFIG['site_name']}"
            response = requests.get(url, headers=self.get_headers(), timeout=30)

            if response.status_code == 200:
                site_data = response.json()
                self.site_id = site_data['id']
                self.log(f"✅ Found SharePoint site ID: {self.site_id}")
                return self.site_id
            else:
                self.log(f"❌ Error getting SharePoint site ID: {response.status_code}")
                return None

        except Exception as e:
            self.log(f"❌ Error getting SharePoint site ID: {str(e)}")
            return None

    def download_excel_file(self):
        """Download Excel file from SharePoint"""
        try:
            self.log(f"📥 Downloading CIP plan file from SharePoint...")

            # Get file download URL using file ID
            url = f"{self.base_url}/sites/{self.get_site_id()}/drive/items/{CIP_PLAN_FILE_ID}"
            response = requests.get(url, headers=self.get_headers(), timeout=30)

            if response.status_code == 200:
                file_info = response.json()
                download_url = file_info.get('@microsoft.graph.downloadUrl')

                if download_url:
                    # Download file content
                    self.log(f"✅ Got download URL, downloading content...")
                    file_response = requests.get(download_url, timeout=60)

                    if file_response.status_code == 200:
                        # Read Excel from memory
                        excel_data = io.BytesIO(file_response.content)
                        self.log(f"✅ Downloaded {len(file_response.content)} bytes")
                        
                        try:
                            excel_file = pd.ExcelFile(excel_data)
                            sheets_data = {}
                            
                            self.log(f"Excel sheets found: {excel_file.sheet_names}")
                            
                            for sheet_name in excel_file.sheet_names:
                                excel_data.seek(0)
                                df = pd.read_excel(excel_data, sheet_name=sheet_name)
                                sheets_data[sheet_name] = df
                                self.log(f"✅ Sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
                            
                            self.log(f"✅ Successfully downloaded CIP plan file")
                            return sheets_data
                            
                        except Exception as e:
                            self.log(f"❌ Error reading Excel file: {str(e)}")
                            return None
                    else:
                        self.log(f"❌ Error downloading file content: {file_response.status_code}")
                else:
                    self.log(f"❌ No download URL found for CIP plan file")
            else:
                self.log(f"❌ Error getting file info: {response.status_code}")

        except Exception as e:
            self.log(f"❌ Error downloading CIP plan file: {str(e)}")

        return None

    def upload_excel_file(self, sheets_data):
        """Upload updated Excel file back to SharePoint with professional formatting"""
        max_retries = 5
        retry_delay = 30  # seconds
        
        try:
            self.log(f"📤 Creating professionally formatted Excel file...")

            # Create formatted Excel file using the new formatting function
            wb = create_formatted_excel(sheets_data)
            
            # Save to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            excel_content = excel_buffer.getvalue()
            
            self.log(f"✅ Created professionally formatted Excel file with {len(excel_content)} bytes")

            # Upload to SharePoint with retry logic
            upload_url = f"{self.base_url}/sites/{self.get_site_id()}/drive/items/{CIP_PLAN_FILE_ID}/content"

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }

            for attempt in range(max_retries):
                try:
                    self.log(f"Upload attempt {attempt + 1}/{max_retries}")
                    
                    response = requests.put(upload_url, headers=headers, data=excel_content, timeout=60)

                    if response.status_code in [200, 201]:
                        self.log(f"✅ Successfully uploaded professionally formatted CIP plan to SharePoint")
                        return True
                    elif response.status_code == 423:
                        # File is locked
                        self.log(f"⚠️ File is locked (attempt {attempt + 1}/{max_retries})")
                        if attempt < max_retries - 1:
                            self.log(f"⏳ Waiting {retry_delay} seconds before retry...")
                            time.sleep(retry_delay)
                            continue
                        else:
                            self.log(f"❌ File remains locked after {max_retries} attempts")
                            # Try to save to a backup location or with different name
                            return self.upload_backup_file(excel_content)
                    elif response.status_code == 401:
                        # Token expired, try refresh
                        self.log("🔄 Token expired during upload, refreshing...")
                        if self.refresh_access_token():
                            self.update_github_secrets()
                            headers['Authorization'] = f'Bearer {self.access_token}'
                            continue
                        else:
                            self.log("❌ Token refresh failed during upload")
                            return False
                    else:
                        self.log(f"❌ Error uploading to SharePoint: {response.status_code}")
                        self.log(f"Response: {response.text[:500]}")
                        if attempt < max_retries - 1:
                            self.log(f"⏳ Retrying in {retry_delay} seconds...")
                            time.sleep(retry_delay)
                            continue
                        else:
                            return False

                except requests.exceptions.RequestException as e:
                    self.log(f"❌ Network error during upload: {str(e)}")
                    if attempt < max_retries - 1:
                        self.log(f"⏳ Retrying in {retry_delay} seconds...")
                        time.sleep(retry_delay)
                        continue
                    else:
                        return False

            return False

        except Exception as e:
            self.log(f"❌ Error uploading to SharePoint: {str(e)}")
            return False

    def upload_backup_file(self, excel_content):
        """Upload to a backup file when original is locked"""
        try:
            # Generate backup filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"CIP_plan_backup_{timestamp}.xlsx"
            
            self.log(f"🔄 Uploading to backup file: {backup_filename}")
            
            # Upload to the same folder but with different name
            # First get the parent folder
            file_info_url = f"{self.base_url}/sites/{self.get_site_id()}/drive/items/{CIP_PLAN_FILE_ID}"
            response = requests.get(file_info_url, headers=self.get_headers(), timeout=30)
            
            if response.status_code == 200:
                file_info = response.json()
                parent_id = file_info.get('parentReference', {}).get('id')
                
                if parent_id:
                    # Upload to parent folder with new name
                    upload_url = f"{self.base_url}/sites/{self.get_site_id()}/drive/items/{parent_id}:/{backup_filename}:/content"
                    
                    headers = {
                        'Authorization': f'Bearer {self.access_token}',
                        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    }
                    
                    response = requests.put(upload_url, headers=headers, data=excel_content, timeout=60)
                    
                    if response.status_code in [200, 201]:
                        self.log(f"✅ Successfully uploaded backup file: {backup_filename}")
                        self.log(f"⚠️ Original file was locked, please check and rename backup file manually")
                        return True
                    else:
                        self.log(f"❌ Failed to upload backup file: {response.status_code}")
                        return False
                else:
                    self.log(f"❌ Could not get parent folder information")
                    return False
            else:
                self.log(f"❌ Could not get file information for backup: {response.status_code}")
                return False
                
        except Exception as e:
            self.log(f"❌ Error uploading backup file: {str(e)}")
            return False

    def update_sheet_data(self, sheet_name, df):
        """Update specific sheet data in SharePoint Excel file"""
        # For now, we'll update the entire file. 
        # In future, could implement partial sheet updates if needed
        pass

# Helper function to parse dates in different formats
def parse_date(date_str):
    """Try to parse date with multiple formats and handle Excel date formats"""
    if not date_str or str(date_str).strip() in ['nan', 'None', '', 'NaT']:
        return None
    
    # If it's already a datetime object, return it
    if isinstance(date_str, datetime):
        return date_str
    
    # If it's a pandas timestamp, convert it
    if hasattr(date_str, 'to_pydatetime'):
        try:
            return date_str.to_pydatetime()
        except:
            pass
    
    # Convert to string and clean
    date_str = str(date_str).strip()
    
    # Handle Excel serial dates (numbers like 45123.0)
    try:
        # If it's a number that could be an Excel date
        if date_str.replace('.', '').isdigit():
            excel_date = float(date_str)
            # Excel date serial numbers are typically > 1 and < 50000 for reasonable dates
            if 1 < excel_date < 50000:
                # Excel epoch is 1900-01-01 (with some quirks)
                excel_epoch = datetime(1900, 1, 1)
                # Excel incorrectly treats 1900 as a leap year, so subtract 2 days
                return excel_epoch + timedelta(days=excel_date - 2)
    except (ValueError, TypeError):
        pass
    
    # Try various date formats
    date_formats = [
        '%B %d, %Y',     # June 7, 2025
        '%d/%m/%Y',      # 07/06/2025
        '%m/%d/%Y',      # 06/07/2025
        '%Y-%m-%d',      # 2025-06-07
        '%d-%m-%Y',      # 07-06-2025
        '%d %B %Y',      # 7 June 2025
        '%d %B, %Y',     # 7 June, 2025
        '%d/%m/%y',      # 07/06/25
        '%m/%d/%y',      # 06/07/25
        '%d-%m-%y',      # 07-06-25
        '%d.%m.%Y',      # 07.06.2025
        '%d.%m.%y',      # 07.06.25
        '%Y/%m/%d',      # 2025/06/07
        '%d-%b-%Y',      # 07-Jun-2025
        '%d-%b-%y',      # 07-Jun-25
        '%d %b %Y',      # 7 Jun 2025
        '%d %b %y',      # 7 Jun 25
        '%Y-%m-%d %H:%M:%S',  # 2025-06-07 00:00:00
        '%B %d %Y',      # June 7 2025 (no comma)
        '%b %d, %Y',     # Jun 7, 2025
        '%b %d %Y',      # Jun 7 2025
    ]
    
    for fmt in date_formats:
        try:
            parsed_date = datetime.strptime(date_str, fmt)
            # Sanity check - date should be between 1900 and 2100
            if 1900 <= parsed_date.year <= 2100:
                return parsed_date
        except (ValueError, TypeError):
            continue
    
    # Try pandas to_datetime as last resort
    try:
        import pandas as pd
        parsed_date = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
        if not pd.isna(parsed_date):
            return parsed_date.to_pydatetime()
    except:
        pass
    
    print(f"⚠️ Warning: Could not parse date: '{date_str}'")
    return None

# Main function to update cleaning schedule using SharePoint
def update_cleaning_schedule():
    global global_processor  # Use global processor
    
    print("Đang cập nhật lịch vệ sinh từ SharePoint...")
    
    # Initialize SharePoint processor
    global_processor = SharePointCIPProcessor()
    
    # Download Excel file from SharePoint
    sheets_data = global_processor.download_excel_file()
    if not sheets_data:
        print("❌ Failed to download CIP plan file")
        return []
    
    # Get or create sheets data
    master_plan_df = sheets_data.get('Master plan', pd.DataFrame())
    cleaning_history_df = sheets_data.get('Cleaning History', pd.DataFrame())
    actual_result_df = sheets_data.get('Actual result', pd.DataFrame())
    
    # Initialize sheets if empty
    if master_plan_df.empty:
        headers = ['Khu vực', 'Thiết bị', 'Phương pháp', 'Tần suất (ngày)', 
                'Ngày vệ sinh gần nhất', 'Ngày kế hoạch vệ sinh tiếp theo', 'Trạng thái', 'Đang chứa sản phẩm']
        master_plan_df = pd.DataFrame(columns=headers)
        sheets_data['Master plan'] = master_plan_df
    
    if cleaning_history_df.empty:
        history_headers = ['Khu vực', 'Thiết bị', 'Phương pháp', 
                        'Tần suất (ngày)', 'Ngày vệ sinh', 'Người thực hiện']
        cleaning_history_df = pd.DataFrame(columns=history_headers)
        sheets_data['Cleaning History'] = cleaning_history_df
    
    if actual_result_df.empty:
        actual_headers = ['Khu vực', 'Thiết bị', 'Phương pháp', 'Tần suất (ngày)', 
                           'Ngày vệ sinh', 'Người thực hiện', 'Kết quả', 'Ghi chú']
        actual_result_df = pd.DataFrame(columns=actual_headers)
        sheets_data['Actual result'] = actual_result_df
    
    # Process Master plan data
    today = datetime.today()
    updated_values = []
    
    # Initialize counters FIRST - CRITICAL!
    processed_count = 0
    status_counts = {'Bình thường': 0, 'Sắp đến hạn': 0, 'Đến hạn': 0, 'Quá hạn': 0, 'Chưa có dữ liệu': 0, 'Lỗi': 0}
    
    # Check if data already has status column - maybe we don't need to calculate
    print(f"🔍 Checking existing status data...")
    
    # Check if there's already a status column with data
    existing_status_col = None
    for col in master_plan_df.columns:
        col_lower = str(col).lower().strip()
        if 'trạng thái' in col_lower:
            existing_status_col = col
            break
    
    if existing_status_col:
        print(f"✅ Found existing status column: '{existing_status_col}'")
        existing_statuses = master_plan_df[existing_status_col].value_counts()
        print(f"📊 Existing status breakdown:")
        for status, count in existing_statuses.items():
            print(f"  - '{status}': {count}")
        
        # If we already have status data, let's use it and just update dates
        use_existing_status = True
        print(f"🔄 Using existing status data instead of recalculating")
    else:
        print(f"⚠️ No existing status column found, will calculate status")
        use_existing_status = False
    
    for idx, row in master_plan_df.iterrows():
        try:
            # Get values, handle missing columns with more flexible column name matching
            area = ''
            device = ''
            method = ''
            freq_str = ''
            last_cleaning = ''
            has_product = ''
            existing_status = ''
            
            # More flexible column matching
            for col in master_plan_df.columns:
                col_lower = str(col).lower().strip()
                if 'khu' in col_lower and 'vực' in col_lower:
                    area = str(row.get(col, '')).strip() if pd.notna(row.get(col, '')) else ''
                elif 'thiết' in col_lower and 'bị' in col_lower:
                    device = str(row.get(col, '')).strip() if pd.notna(row.get(col, '')) else ''
                elif 'phương' in col_lower and 'pháp' in col_lower:
                    method = str(row.get(col, '')).strip() if pd.notna(row.get(col, '')) else ''
                elif 'tần' in col_lower and 'suất' in col_lower:
                    freq_str = str(row.get(col, '')).strip() if pd.notna(row.get(col, '')) else ''
                elif 'ngày' in col_lower and 'vệ sinh' in col_lower and 'gần' in col_lower:
                    last_cleaning = str(row.get(col, '')).strip() if pd.notna(row.get(col, '')) else ''
                elif 'chứa' in col_lower and 'sản phẩm' in col_lower:
                    has_product = str(row.get(col, '')).strip() if pd.notna(row.get(col, '')) else ''
                elif 'trạng thái' in col_lower:
                    existing_status = str(row.get(col, '')).strip() if pd.notna(row.get(col, '')) else ''
            
            # Debug first few rows
            if idx < 5:
                print(f"🔍 Row {idx} extracted data:")
                print(f"  area: '{area}', device: '{device}', method: '{method}'")
                print(f"  freq_str: '{freq_str}', last_cleaning: '{last_cleaning}', has_product: '{has_product}'")
                print(f"  existing_status: '{existing_status}'")
            
            # Skip empty rows
            if not area and not device:
                if idx < 5:
                    print(f"  -> Skipping empty row {idx}")
                continue
            
            # If we have existing status and it's valid, use it
            if use_existing_status and existing_status and existing_status not in ['nan', 'None', '']:
                current_status = existing_status
                
                # Still try to calculate next plan date if we have the data
                next_plan_str = ''
                if last_cleaning and last_cleaning not in ['nan', 'None', '']:
                    if freq_str and freq_str not in ['nan', 'None', '']:
                        try:
                            freq = int(float(freq_str))
                            last_cleaning_date = parse_date(last_cleaning)
                            if last_cleaning_date:
                                next_plan_date = last_cleaning_date + timedelta(days=freq)
                                next_plan_str = next_plan_date.strftime('%d/%m/%Y')
                        except (ValueError, TypeError):
                            pass
                
                updated_values.append([area, device, method, freq_str, last_cleaning, next_plan_str, current_status, has_product])
                
                # Safe increment using .get() method
                status_counts[current_status] = status_counts.get(current_status, 0) + 1
                processed_count += 1
                
                if idx < 5:
                    print(f"  -> Using existing status: {current_status}")
                continue
                
            # Otherwise, calculate status as before
            if not last_cleaning or last_cleaning in ['nan', 'None', '']:
                status = "Chưa có dữ liệu"
                updated_values.append([area, device, method, freq_str, last_cleaning, "", status, has_product])
                status_counts[status] = status_counts.get(status, 0) + 1
                if idx < 5:
                    print(f"  -> Status: {status} (no cleaning date)")
                continue
    
            freq = 0
            if freq_str and freq_str not in ['nan', 'None', '']:
                try:
                    freq = int(float(freq_str))
                except ValueError:
                    freq = 0
    
            last_cleaning_date = parse_date(last_cleaning)
            if not last_cleaning_date:
                status = "Định dạng ngày không hợp lệ"
                updated_values.append([area, device, method, freq_str, last_cleaning, "", status, has_product])
                status_counts['Lỗi'] = status_counts.get('Lỗi', 0) + 1
                if idx < 5:
                    print(f"  -> Status: {status} (invalid date: '{last_cleaning}')")
                continue
    
            next_plan_date = last_cleaning_date + timedelta(days=freq)
            next_plan_str = next_plan_date.strftime('%d/%m/%Y')
    
            days_until_next = (next_plan_date.date() - today.date()).days
            
            if days_until_next > 7:
                current_status = 'Bình thường'
            elif days_until_next > 0:
                current_status = 'Sắp đến hạn'
            elif days_until_next == 0:
                current_status = 'Đến hạn'
            else:
                current_status = 'Quá hạn'
    
            updated_values.append([area, device, method, freq_str, last_cleaning, next_plan_str, current_status, has_product])
            status_counts[current_status] = status_counts.get(current_status, 0) + 1
            processed_count += 1
            
            # Debug first few rows
            if idx < 5:
                print(f"  -> Calculated status: {current_status} (days until next: {days_until_next})")
                print(f"  -> Last cleaning: {last_cleaning_date.strftime('%d/%m/%Y')}, Next: {next_plan_str}")
            
            # Update the DataFrame (only if we calculated new values)
            if not use_existing_status:
                # Find the correct column names for updating
                for col in master_plan_df.columns:
                    col_lower = str(col).lower().strip()
                    if 'kế hoạch' in col_lower or ('ngày' in col_lower and 'tiếp theo' in col_lower):
                        master_plan_df.at[idx, col] = next_plan_str
                    elif 'trạng thái' in col_lower:
                        master_plan_df.at[idx, col] = current_status
            
        except Exception as e:
            print(f"❌ Error processing row {idx}: {str(e)}")
            print(f"❌ Full traceback: {traceback.format_exc()}")
            status_counts['Lỗi'] = status_counts.get('Lỗi', 0) + 1
            continue
    
    # Print processing summary
    print(f"\n📊 Processing Summary:")
    print(f"  - Total rows processed: {processed_count}")
    print(f"  - Status breakdown:")
    for status, count in status_counts.items():
        if count > 0:
            print(f"    - {status}: {count}")
    
    print(f"\n🎯 Due/Overdue Equipment Check:")
    due_count = status_counts.get('Đến hạn', 0) + status_counts.get('Quá hạn', 0)
    print(f"  - Equipment due for cleaning: {due_count}")
    print(f"    - Đến hạn: {status_counts.get('Đến hạn', 0)}")
    print(f"    - Quá hạn: {status_counts.get('Quá hạn', 0)}")
    
    # Update sheets data
    sheets_data['Master plan'] = master_plan_df
    
    # Update Actual Result with new cleaning records
    print("Kiểm tra và cập nhật bản ghi vệ sinh mới...")
    
    # Read existing records from Actual Result
    existing_records = set()  # Set of unique cleaning records (device + date)
    
    for idx, row in actual_result_df.iterrows():
        device_name = str(row.get('Thiết bị', '')).strip() if pd.notna(row.get('Thiết bị', '')) else ''
        cleaning_date_str = str(row.get('Ngày vệ sinh', '')).strip() if pd.notna(row.get('Ngày vệ sinh', '')) else ''
        if device_name and cleaning_date_str:
            record_key = f"{device_name}_{cleaning_date_str}"
            existing_records.add(record_key)
    
    # Identify new cleaning records from Master plan
    new_cleaning_records = []
    
    for row in updated_values:
        area, device, method, freq_str, last_cleaning, next_plan_str, status, has_product = row
        
        # Skip if no cleaning date or format is invalid
        if not last_cleaning or "không hợp lệ" in status.lower() or "chưa có dữ liệu" in status.lower():
            continue
            
        # Create unique key for this cleaning record
        record_key = f"{device}_{last_cleaning}"
        
        # Add to Actual Result if not already recorded
        if record_key not in existing_records:
            # Default values for new records
            person = "Tự động"  # Placeholder or default person
            result = "Đạt"      # Default result
            notes = ""          # Empty notes
            
            # Add new cleaning record
            new_cleaning_records.append({
                'Khu vực': area,
                'Thiết bị': device,
                'Phương pháp': method,
                'Tần suất (ngày)': freq_str,
                'Ngày vệ sinh': last_cleaning,
                'Người thực hiện': person,
                'Kết quả': result,
                'Ghi chú': notes
            })
            
            # Mark as processed to avoid duplicates
            existing_records.add(record_key)
    
    # Add new cleaning records to Actual Result sheet
    if new_cleaning_records:
        new_df = pd.DataFrame(new_cleaning_records)
        actual_result_df = pd.concat([actual_result_df, new_df], ignore_index=True)
        sheets_data['Actual result'] = actual_result_df
        print(f"Đã thêm {len(new_cleaning_records)} bản ghi vệ sinh mới vào Actual Result")
    else:
        print("Không có bản ghi vệ sinh mới để thêm vào Actual Result")
    
    print(f"Đã cập nhật {len(updated_values)} thiết bị.")
    
    # Try to upload updated file back to SharePoint
    upload_success = False
    if len(updated_values) > 0:
        print(f"\n📤 Attempting to upload updated file...")
        try:
            upload_success = global_processor.upload_excel_file(sheets_data)
        except Exception as e:
            print(f"⚠️ Upload failed with error: {str(e)}")
            upload_success = False
    
    # Create local backup if upload failed but we have data
    if not upload_success and len(updated_values) > 0:
        try:
            backup_filename = f"CIP_plan_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = create_formatted_excel(sheets_data)
            wb.save(backup_filename)
            print(f"💾 Created local backup: {backup_filename}")
        except Exception as e:
            print(f"❌ Failed to create local backup: {str(e)}")
    
    return updated_values

# Function to add a new cleaning record
def add_cleaning_record(area, device, method, freq, cleaning_date, person, result="Đạt", notes=""):
    """
    Add a new cleaning record and update Master plan and Actual Result
    Note: This function would need SharePoint integration for direct updates
    """
    print(f"Adding cleaning record for {device} on {cleaning_date}")
    # Implementation would require SharePoint integration
    return "Thành công"

# Function to update cleaning result
def update_cleaning_result(device, cleaning_date, result, notes=""):
    """
    Update the result of a cleaning record in the Actual Result sheet
    Note: This function would need SharePoint integration for direct updates
    """
    print(f"Updating cleaning result for {device} on {cleaning_date}")
    # Implementation would require SharePoint integration
    return "Thành công"

# Function to update product status
def update_product_status(device, has_product):
    """
    Update the product status for a device in the Master plan
    Note: This function would need SharePoint integration for direct updates
    """
    print(f"Updating product status for {device}")
    # Implementation would require SharePoint integration
    return "Thành công"

# Function to create status chart
def create_status_chart(updated_values):
    try:
        # Create DataFrame for visualization
        df = pd.DataFrame(updated_values, columns=[
            'Khu vực', 'Thiết bị', 'Phương pháp', 'Tần suất (ngày)',
            'Ngày vệ sinh gần nhất', 'Ngày kế hoạch vệ sinh tiếp theo', 'Trạng thái', 'Đang chứa sản phẩm'
        ])
        
        # Set up figure with 2 subplots
        plt.style.use('default')  # Use default style for professional look
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))
        
        # First subplot: Count statuses
        status_counts = df['Trạng thái'].value_counts()
        status_order = ['Bình thường', 'Sắp đến hạn', 'Đến hạn', 'Quá hạn']
        
        # Create a Series with all possible statuses and fill missing with 0
        status_data = pd.Series([0, 0, 0, 0], index=status_order)
        
        # Update with actual counts
        for status, count in status_counts.items():
            if status in status_data.index:
                status_data[status] = count
        
        # Create a bar chart for cleaning status with professional colors
        colors = ['#2D5016', '#7F6000', '#9C6500', '#9C0006']  # Matching Excel colors
        bars = ax1.bar(status_data.index, status_data.values, color=colors)
        ax1.set_title('📊 Thống kê trạng thái thiết bị vệ sinh', fontsize=14, fontweight='bold', pad=20)
        ax1.set_ylabel('Số lượng thiết bị', fontsize=12)
        ax1.grid(axis='y', linestyle='--', alpha=0.3)
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax1.text(bar.get_x() + bar.get_width()/2., height,
                        f'{int(height)}',
                        ha='center', va='bottom', fontweight='bold')
        
        # Rotate x-axis labels if needed
        ax1.tick_params(axis='x', rotation=45)
        
        # Second subplot: Count product status for overdue equipment
        overdue_df = df[df['Trạng thái'].isin(['Đến hạn', 'Quá hạn'])]
        
        if len(overdue_df) > 0:
            # Count devices with/without product
            product_status = overdue_df['Đang chứa sản phẩm'].fillna('').map(
                lambda x: 'Có sản phẩm' if str(x).strip() and str(x).strip() not in ['nan', 'None'] else 'Trống'
            )
            product_counts = product_status.value_counts()
            
            # Ensure both categories are present
            product_data = pd.Series([0, 0], index=['Có sản phẩm', 'Trống'])
            for status, count in product_counts.items():
                if status in product_data.index:
                    product_data[status] = count
            
            # Create a pie chart for product status
            pie_colors = ['#FFCCCB', '#D4F4DD']  # Red for with product, green for empty
            wedges, texts, autotexts = ax2.pie(
                product_data.values,
                labels=product_data.index,
                colors=pie_colors,
                autopct=lambda pct: f'{pct:.1f}%\n({int(pct/100*sum(product_data.values))} thiết bị)' if pct > 0 else '',
                startangle=90,
                textprops={'fontsize': 10, 'fontweight': 'bold'}
            )
            ax2.set_title('🏭 Trạng thái sản phẩm\n(Thiết bị cần vệ sinh)', fontsize=14, fontweight='bold', pad=20)
        else:
            ax2.text(0.5, 0.5, 'Không có thiết bị\ncần vệ sinh', 
                    ha='center', va='center', fontsize=14, fontweight='bold',
                    transform=ax2.transAxes)
            ax2.set_title('🏭 Trạng thái sản phẩm\n(Thiết bị cần vệ sinh)', fontsize=14, fontweight='bold', pad=20)
        
        # Add overall title and footer
        fig.suptitle(f'📈 BÁO CÁO TÌNH TRẠNG VỆ SINH THIẾT BỊ\n{datetime.now().strftime("%d/%m/%Y %H:%M")}', 
                    fontsize=16, fontweight='bold', y=0.95)
        
        plt.tight_layout()
        plt.subplots_adjust(top=0.85)  # Make room for suptitle
        
        # Save chart for email
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', 
                   facecolor='white', edgecolor='none')
        img_buffer.seek(0)
        
        plt.close()  # Close the plot to avoid warnings
        return img_buffer
    
    except Exception as e:
        print(f"Lỗi khi tạo biểu đồ: {str(e)}")
        return None

# Function to create results analysis chart
def create_results_chart():
    try:
        # This would need to get data from SharePoint
        # For now, return None
        return None
    
    except Exception as e:
        print(f"Lỗi khi tạo biểu đồ kết quả: {str(e)}")
        return None

# Modified send_email_report function with Outlook SMTP
def send_email_report(updated_values):
    print("Đang chuẩn bị gửi email báo cáo...")
    
    # Debug: Print all updated values to understand the data structure
    print(f"🔍 Total updated_values: {len(updated_values)}")
    if updated_values:
        print(f"🔍 Sample updated_values (first 3):")
        for i, row in enumerate(updated_values[:3]):
            print(f"  Row {i}: {row}")
            if len(row) > 6:
                print(f"    Status (index 6): '{row[6]}'")
    
    # Filter devices requiring attention
    due_rows = [row for row in updated_values if len(row) > 6 and row[6] in ['Đến hạn', 'Quá hạn']]
    
    print(f"🔍 Filtering logic:")
    print(f"  - Looking for status in ['Đến hạn', 'Quá hạn']")
    print(f"  - Found {len(due_rows)} due/overdue devices")
    
    # Debug: Print status breakdown
    status_breakdown = {}
    for row in updated_values:
        if len(row) > 6:
            status = row[6]
            status_breakdown[status] = status_breakdown.get(status, 0) + 1
    
    print(f"🔍 Status breakdown from updated_values:")
    for status, count in status_breakdown.items():
        print(f"  - '{status}': {count}")
    
    if due_rows:
        print(f"✅ Found {len(due_rows)} devices requiring attention")
        
        # Debug: Print due devices
        print(f"🔍 Due devices details:")
        for i, row in enumerate(due_rows[:5]):  # Show first 5
            print(f"  {i+1}. {row[0]} - {row[1]} - Status: {row[6]}")
        
        try:
            # Create charts
            status_img_buffer = create_status_chart(updated_values)
            results_img_buffer = create_results_chart()
            
            # Split the devices by area
            ro_station_rows = [row for row in due_rows if 'trạm ro' in str(row[0]).lower()]
            other_area_rows = [row for row in due_rows if 'trạm ro' not in str(row[0]).lower()]
            
            print(f"🔍 Area breakdown:")
            print(f"  - RO station devices: {len(ro_station_rows)}")
            print(f"  - Other area devices: {len(other_area_rows)}")
            
            # Define email recipient lists
            ro_recipients = [
                "mmb-ktcncsd@msc.masangroup.com", 
                "mmb-baotri-utilities@msc.masangroup.com", 
            ]
            
            other_recipients = [
                "mmb-ktcncsd@msc.masangroup.com",
            ]
            
            # Send RO station email if there are relevant items
            if ro_station_rows:
                print(f"📧 Sending email for RO station ({len(ro_station_rows)} devices)")
                send_area_specific_email(
                    ro_station_rows, 
                    ro_recipients, 
                    "Trạm RO", 
                    status_img_buffer, 
                    results_img_buffer
                )
            
            # Send other areas email if there are relevant items
            if other_area_rows:
                print(f"📧 Sending email for other areas ({len(other_area_rows)} devices)")
                send_area_specific_email(
                    other_area_rows, 
                    other_recipients, 
                    "Khu vực muối, cốt, chế biến mắm", 
                    status_img_buffer, 
                    results_img_buffer
                )
                
            print("✅ Email đã được gửi kèm bảng HTML và biểu đồ.")
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi gửi email: {str(e)}")
            print(f"❌ Traceback: {traceback.format_exc()}")
            return False
    else:
        print("⚠️ Không có thiết bị đến hạn/quá hạn, không gửi email.")
        print("🔍 This might be due to:")
        print("  1. Date parsing issues")
        print("  2. Incorrect status calculation")
        print("  3. Data structure problems")
        print("  4. Column mapping issues")
        return True

# Helper function to send area-specific emails with Graph API
def send_area_specific_email(filtered_rows, recipients, area_name, status_img_buffer, results_img_buffer):
    """
    Send an email for a specific area with the filtered rows using Microsoft Graph API
    """
    global global_processor  # Use the global processor
    
    try:
        if not global_processor or not global_processor.access_token:
            print("❌ No valid access token for Graph API")
            return False
            
        print(f"📧 Preparing email via Microsoft Graph API for {area_name}")
        
        # Prepare data for email summary
        empty_tanks = [row for row in filtered_rows if not str(row[7]).strip() or str(row[7]).strip() in ['nan', 'None']]
        filled_tanks = [row for row in filtered_rows if str(row[7]).strip() and str(row[7]).strip() not in ['nan', 'None']]
        
        # Create HTML content with improved styling
        html_content = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ 
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                    line-height: 1.6;
                    color: #333;
                    background-color: #f8f9fa;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background-color: white;
                    padding: 20px;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                .header {{
                    background: linear-gradient(135deg, #366092, #4a7bb7);
                    color: white;
                    padding: 20px;
                    border-radius: 8px 8px 0 0;
                    margin: -20px -20px 20px -20px;
                    text-align: center;
                }}
                .header h1 {{
                    margin: 0;
                    font-size: 24px;
                    font-weight: bold;
                }}
                .summary {{
                    background-color: #f8f9fa;
                    padding: 20px;
                    border-radius: 8px;
                    margin: 20px 0;
                    border-left: 4px solid #366092;
                }}
                .summary h3 {{
                    color: #366092;
                    margin-top: 0;
                    font-size: 18px;
                }}
                .kpi {{
                    display: inline-block;
                    background: white;
                    padding: 15px;
                    margin: 10px;
                    border-radius: 8px;
                    text-align: center;
                    box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                    min-width: 150px;
                }}
                .kpi-value {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #C00000;
                }}
                .kpi-label {{
                    font-size: 12px;
                    color: #666;
                    margin-top: 5px;
                }}
                table {{ 
                    border-collapse: collapse; 
                    width: 100%; 
                    margin-top: 20px;
                    font-size: 11px;
                }}
                th, td {{ 
                    border: 1px solid #ddd; 
                    padding: 12px 8px; 
                    text-align: left; 
                    vertical-align: top;
                }}
                th {{ 
                    background: linear-gradient(135deg, #366092, #4a7bb7);
                    color: white;
                    font-weight: bold;
                    text-align: center;
                    font-size: 10px;
                }}
                .overdue {{ background-color: #ffebee; border-left: 4px solid #f44336; }}
                .due-today {{ background-color: #fff8e1; border-left: 4px solid #ff9800; }}
                .has-product {{ 
                    color: #C00000; 
                    font-weight: bold;
                    background-color: #ffebee;
                    padding: 4px 8px;
                    border-radius: 4px;
                }}
                .empty {{ 
                    color: #2e7d32;
                    background-color: #e8f5e8;
                    padding: 4px 8px;
                    border-radius: 4px;
                }}
                .status-normal {{ background-color: #e8f5e8; color: #2e7d32; }}
                .status-coming {{ background-color: #fff8e1; color: #f57c00; }}
                .status-due {{ background-color: #fff3e0; color: #f57c00; font-weight: bold; }}
                .status-overdue {{ background-color: #ffebee; color: #c62828; font-weight: bold; }}
                .footer {{ 
                    margin-top: 30px; 
                    padding-top: 20px;
                    border-top: 1px solid #e0e0e0;
                    font-size: 12px; 
                    color: #666; 
                    text-align: center;
                }}
                .priority-high {{
                    background-color: #ffcdd2;
                    border-left: 5px solid #f44336;
                }}
                .alert-box {{
                    background-color: #ffebee;
                    border: 1px solid #f44336;
                    border-radius: 8px;
                    padding: 15px;
                    margin: 20px 0;
                }}
                .alert-title {{
                    color: #c62828;
                    font-weight: bold;
                    font-size: 16px;
                    margin-bottom: 10px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🏭 BÁO CÁO VỆ SINH THIẾT BỊ</h1>
                    <p style="margin: 10px 0 0 0; font-size: 16px;">{area_name} - {datetime.today().strftime("%d/%m/%Y")}</p>
                </div>
                
                <div class="summary">
                    <h3>📊 TỔNG QUAN TÌNH TRẠNG</h3>
                    <div style="text-align: center;">
                        <div class="kpi">
                            <div class="kpi-value">{len(filtered_rows)}</div>
                            <div class="kpi-label">Tổng thiết bị cần vệ sinh</div>
                        </div>
                        <div class="kpi">
                            <div class="kpi-value" style="color: #2e7d32;">{len(empty_tanks)}</div>
                            <div class="kpi-label">Thiết bị trống<br>(Có thể vệ sinh ngay)</div>
                        </div>
                        <div class="kpi">
                            <div class="kpi-value" style="color: #f57c00;">{len(filled_tanks)}</div>
                            <div class="kpi-label">Thiết bị chứa sản phẩm<br>(Cần lên kế hoạch)</div>
                        </div>
                    </div>
                </div>
        """
        
        # Add alert for critical equipment
        critical_count = len([row for row in filtered_rows if row[6] == 'Quá hạn' and str(row[7]).strip() and str(row[7]).strip() not in ['nan', 'None']])
        if critical_count > 0:
            html_content += f"""
                <div class="alert-box">
                    <div class="alert-title">🚨 CẢNH BÁO KHẨN CẤP</div>
                    <p><strong>{critical_count} thiết bị</strong> đã quá hạn vệ sinh và đang chứa sản phẩm. Cần xử lý ngay lập tức để đảm bảo chất lượng sản phẩm!</p>
                </div>
            """
        
        html_content += """
                <h3>📋 DANH SÁCH CHI TIẾT THIẾT BỊ CẦN VỆ SINH</h3>
                <table>
                    <thead>
                        <tr>
                            <th style="width: 12%;">Khu vực</th>
                            <th style="width: 18%;">Thiết bị</th>
                            <th style="width: 10%;">Phương pháp</th>
                            <th style="width: 8%;">Tần suất<br>(ngày)</th>
                            <th style="width: 12%;">Ngày vệ sinh<br>gần nhất</th>
                            <th style="width: 12%;">Ngày kế hoạch<br>vệ sinh tiếp theo</th>
                            <th style="width: 10%;">Trạng thái</th>
                            <th style="width: 18%;">Tình trạng<br>sản phẩm</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        # Sort rows: empty tanks first, then by status priority
        def sort_priority(row):
            area, device, method, freq_str, last_cleaning, next_plan_str, status, has_product = row
            # Priority: empty tanks with overdue status first
            if status == "Quá hạn":
                if not str(has_product).strip() or str(has_product).strip() in ['nan', 'None']:
                    return 0  # Highest priority: overdue + empty
                else:
                    return 1  # Second priority: overdue + has product
            elif status == "Đến hạn":
                if not str(has_product).strip() or str(has_product).strip() in ['nan', 'None']:
                    return 2  # Third priority: due + empty
                else:
                    return 3  # Fourth priority: due + has product
            return 4
        
        sorted_rows = sorted(filtered_rows, key=sort_priority)
        
        for row in sorted_rows:
            area, device, method, freq_str, last_cleaning, next_plan_str, status, has_product = row
            
            # Define CSS class based on status
            css_class = ""
            status_class = ""
            if status == "Quá hạn":
                css_class = "overdue"
                status_class = "status-overdue"
            elif status == "Đến hạn":
                css_class = "due-today"
                status_class = "status-due"
            
            # Check if this is a critical combination
            is_critical = (status == "Quá hạn" and str(has_product).strip() and str(has_product).strip() not in ['nan', 'None'])
            if is_critical:
                css_class += " priority-high"
            
            # Define product status class and display
            has_product_clean = str(has_product).strip()
            if has_product_clean and has_product_clean not in ['nan', 'None']:
                product_class = "has-product"
                product_display = "🔴 Có sản phẩm"
                if is_critical:
                    product_display = "🚨 CÓ SẢN PHẨM (KHẨN CẤP)"
            else:
                product_class = "empty"
                product_display = "✅ Trống"
            
            html_content += f"""
                    <tr class="{css_class}">
                        <td style="font-weight: bold;">{area}</td>
                        <td style="font-weight: bold;">{device}</td>
                        <td>{method}</td>
                        <td style="text-align: center;">{freq_str}</td>
                        <td style="text-align: center;">{last_cleaning}</td>
                        <td style="text-align: center;">{next_plan_str}</td>
                        <td class="{status_class}" style="text-align: center; font-weight: bold;">{status}</td>
                        <td class="{product_class}" style="text-align: center;">{product_display}</td>
                    </tr>
            """
        
        html_content += f"""
                </tbody>
            </table>
            
            <div class="footer">
                <p><strong>📍 Hướng dẫn xử lý:</strong></p>
                <p>🟢 <strong>Thiết bị trống:</strong> Có thể tiến hành vệ sinh ngay lập tức</p>
                <p>🟡 <strong>Thiết bị có sản phẩm:</strong> Cần lên kế hoạch vệ sinh sau khi xử lý sản phẩm</p>
                <p>🔴 <strong>Thiết bị quá hạn + có sản phẩm:</strong> Ưu tiên cao nhất, xử lý ngay</p>
                <br>
                <p>📂 Vui lòng truy cập SharePoint để cập nhật trạng thái sau khi hoàn thành vệ sinh.</p>
                <p>🤖 Email này được tự động tạo bởi hệ thống CIP Management. Vui lòng không trả lời.</p>
                <p>🕒 Thời gian tạo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
            </div>
            
            </div>
        </body>
        </html>
        """
        
        # Prepare email data for Graph API
        email_data = {
            "message": {
                "subject": f"🏭 Báo cáo vệ sinh thiết bị - {area_name} - {datetime.today().strftime('%d/%m/%Y')} ({'🚨 KHẨN CẤP' if critical_count > 0 else '📋 Thông thường'})",
                "body": {
                    "contentType": "HTML",
                    "content": html_content
                },
                "toRecipients": []
            }
        }
        
        # Add recipients
        for recipient in recipients:
            email_data["message"]["toRecipients"].append({
                "emailAddress": {
                    "address": recipient
                }
            })
        
        # Prepare attachments if available
        attachments = []
        
        if status_img_buffer:
            status_img_buffer.seek(0)
            status_img_data = status_img_buffer.read()
            status_img_b64 = base64.b64encode(status_img_data).decode('utf-8')
            
            attachments.append({
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": f"cleaning_status_chart_{datetime.now().strftime('%Y%m%d')}.png",
                "contentType": "image/png",
                "contentBytes": status_img_b64
            })
        
        if results_img_buffer:
            results_img_buffer.seek(0)
            results_img_data = results_img_buffer.read()
            results_img_b64 = base64.b64encode(results_img_data).decode('utf-8')
            
            attachments.append({
                "@odata.type": "#microsoft.graph.fileAttachment", 
                "name": f"cleaning_results_chart_{datetime.now().strftime('%Y%m%d')}.png",
                "contentType": "image/png",
                "contentBytes": results_img_b64
            })
        
        if attachments:
            email_data["message"]["attachments"] = attachments
        
        # Send email via Graph API
        graph_url = "https://graph.microsoft.com/v1.0/me/sendMail"
        headers = {
            'Authorization': f'Bearer {global_processor.access_token}',
            'Content-Type': 'application/json'
        }
        
        print(f"📤 Sending email via Graph API to {len(recipients)} recipients...")
        print(f"🔗 Graph URL: {graph_url}")
        
        response = requests.post(graph_url, headers=headers, json=email_data, timeout=60)
        
        if response.status_code == 202:
            print("✅ Email sent successfully via Graph API")
            print(f"✅ Email cho {area_name} đã được gửi đến {len(recipients)} người nhận.")
            return True
        elif response.status_code == 401:
            print("❌ Graph API Authentication Error - Token may have expired")
            print("🔄 Attempting to refresh token...")
            if global_processor.refresh_access_token():
                print("✅ Token refreshed, retrying email send...")
                headers['Authorization'] = f'Bearer {global_processor.access_token}'
                response = requests.post(graph_url, headers=headers, json=email_data, timeout=60)
                if response.status_code == 202:
                    print("✅ Email sent successfully after token refresh")
                    return True
            print("❌ Failed to send email even after token refresh")
            return False
        elif response.status_code == 403:
            print("❌ Graph API Permission Error")
            print("💡 Please ensure Mail.Send permission is granted in Azure App Registration:")
            print("   1. Go to Azure Portal → App registrations")
            print("   2. Find your app → API permissions")
            print("   3. Add Microsoft Graph → Delegated permissions → Mail.Send")
            print("   4. Grant admin consent")
            return False
        else:
            print(f"❌ Graph API Error: {response.status_code}")
            print(f"❌ Response: {response.text[:500]}")
            return False
            
    except Exception as e:
        print(f"❌ Error sending email via Graph API: {str(e)}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        return False

# Main function to run everything
def run_update():
    print("Bắt đầu cập nhật hệ thống vệ sinh thiết bị từ SharePoint...")
    
    try:
        # Update cleaning schedule and get updated values
        updated_values = update_cleaning_schedule()
        
        # Send email report
        send_email_report(updated_values)
        
        print("Hoàn thành cập nhật.")
        return True
    except Exception as e:
        print(f"Lỗi: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return False

# Additional utility functions for testing and maintenance

def test_excel_formatting():
    """
    Test function to create a sample Excel file with formatting
    """
    print("🧪 Testing Excel formatting...")
    
    # Create sample data
    sample_data = {
        'Master plan': pd.DataFrame({
            'Khu vực': ['Lọc thô', 'Lọc thô', 'Nấng - hạ', 'Lọc KB/ tủ', 'Đường ôn'] * 4,
            'Thiết bị': [f'Bồn {i}' for i in range(1, 21)],
            'Phương pháp': ['CIP 1', 'CIP 2'] * 10,
            'Tần suất (ngày)': [7, 15, 30, 60] * 5,
            'Ngày vệ sinh gần nhất': [
                '10/06/2025', '25/04/2025', '09/06/2025', '09/06/2025', '22/06/2025',
                '27/07/2025', '12/07/2025', '27/07/2025', '13/07/2025', '20/07/2025',
                '20/07/2025', '26/07/2025', '26/07/2025', '20/07/2025', '20/07/2025',
                '25/07/2025', '28/07/2025', '30/07/2025', '01/08/2025', '05/08/2025'
            ],
            'Ngày kế hoạch vệ sinh tiếp theo': [
                '17/06/2025', '10/05/2025', '09/07/2025', '08/08/2025', '22/07/2025',
                '03/08/2025', '27/07/2025', '03/08/2025', '28/07/2025', '04/08/2025',
                '04/08/2025', '10/08/2025', '10/08/2025', '04/08/2025', '04/08/2025',
                '09/08/2025', '12/08/2025', '14/08/2025', '16/08/2025', '20/08/2025'
            ],
            'Trạng thái': [
                'Quá hạn', 'Quá hạn', 'Quá hạn', 'Quá hạn', 'Bình thường',
                'Bình thường', 'Đến hạn', 'Bình thường', 'Đến hạn', 'Bình thường',
                'Bình thường', 'Bình thường', 'Bình thường', 'Bình thường', 'Bình thường',
                'Bình thường', 'Bình thường', 'Bình thường', 'Bình thường', 'Bình thường'
            ],
            'Đang chứa sản phẩm': ['Quá hạn', '', 'Quá hạn', 'Quá hạn', ''] * 4
        }),
        
        'Cleaning History': pd.DataFrame({
            'Khu vực': ['Lọc thô', 'Nấng - hạ', 'Đường ôn'] * 3,
            'Thiết bị': [f'Bồn {i}' for i in range(1, 10)],
            'Phương pháp': ['CIP 1', 'CIP 2'] * 5 + ['CIP 1'],
            'Tần suất (ngày)': [7, 15, 30] * 3,
            'Ngày vệ sinh': [
                '20/07/2025', '21/07/2025', '22/07/2025', '23/07/2025', '24/07/2025',
                '25/07/2025', '26/07/2025', '27/07/2025', '28/07/2025'
            ],
            'Người thực hiện': ['Nguyễn A', 'Trần B', 'Lê C'] * 3
        }),
        
        'Actual result': pd.DataFrame({
            'Khu vực': ['Lọc thô', 'Nấng - hạ', 'Đường ôn'] * 2,
            'Thiết bị': [f'Bồn {i}' for i in range(1, 7)],
            'Phương pháp': ['CIP 1', 'CIP 2'] * 3,
            'Tần suất (ngày)': [7, 15, 30] * 2,
            'Ngày vệ sinh': [
                '20/07/2025', '21/07/2025', '22/07/2025', 
                '23/07/2025', '24/07/2025', '25/07/2025'
            ],
            'Người thực hiện': ['Nguyễn A', 'Trần B'] * 3,
            'Kết quả': ['Đạt', 'Đạt', 'Không đạt', 'Đạt', 'Đạt', 'Đạt'],
            'Ghi chú': ['', '', 'Cần làm lại', '', '', '']
        })
    }
    
    # Create formatted Excel file
    wb = create_formatted_excel(sample_data)
    
    # Save test file
    test_filename = f"CIP_Plan_Test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(test_filename)
    
    print(f"✅ Created test file: {test_filename}")
    print("📋 File features:")
    print("  - Professional color coding by status")
    print("  - Auto-adjusted column widths")
    print("  - Consistent date formatting (DD/MM/YYYY)")
    print("  - Freeze panes and auto-filters")
    print("  - Summary statistics section")
    print("  - Critical equipment highlighting")
    
    return test_filename

def create_local_backup(sheets_data, filename_suffix="manual"):
    """
    Create a local backup of the Excel file with professional formatting
    """
    try:
        backup_filename = f"CIP_plan_backup_{filename_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = create_formatted_excel(sheets_data)
        wb.save(backup_filename)
        print(f"💾 Created local backup: {backup_filename}")
        return backup_filename
    except Exception as e:
        print(f"❌ Failed to create local backup: {str(e)}")
        return None

def validate_data_integrity(sheets_data):
    """
    Validate data integrity and consistency across sheets
    """
    print("🔍 Validating data integrity...")
    
    issues = []
    
    # Check Master plan
    if 'Master plan' in sheets_data:
        master_df = sheets_data['Master plan']
        
        # Check for missing required columns
        required_columns = ['Khu vực', 'Thiết bị', 'Phương pháp', 'Tần suất (ngày)', 'Trạng thái']
        for col in required_columns:
            if col not in master_df.columns:
                issues.append(f"Missing required column in Master plan: {col}")
        
        # Check for empty critical fields
        if 'Thiết bị' in master_df.columns:
            empty_devices = master_df['Thiết bị'].isna().sum()
            if empty_devices > 0:
                issues.append(f"Found {empty_devices} rows with empty device names")
        
        # Check date formats
        date_columns = [col for col in master_df.columns if 'ngày' in col.lower()]
        for date_col in date_columns:
            invalid_dates = 0
            for idx, date_val in master_df[date_col].items():
                if pd.notna(date_val) and str(date_val).strip():
                    if not parse_date(str(date_val)):
                        invalid_dates += 1
            
            if invalid_dates > 0:
                issues.append(f"Found {invalid_dates} invalid dates in column {date_col}")
    
    # Check consistency between sheets
    if 'Master plan' in sheets_data and 'Actual result' in sheets_data:
        master_devices = set(sheets_data['Master plan']['Thiết bị'].dropna())
        actual_devices = set(sheets_data['Actual result']['Thiết bị'].dropna())
        
        # Devices in Actual result but not in Master plan
        orphaned_devices = actual_devices - master_devices
        if orphaned_devices:
            issues.append(f"Found {len(orphaned_devices)} devices in Actual result not in Master plan")
    
    # Print results
    if issues:
        print("⚠️ Data integrity issues found:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("✅ Data integrity validation passed")
    
    return len(issues) == 0

def generate_compliance_report(updated_values):
    """
    Generate a detailed compliance report
    """
    if not updated_values:
        print("❌ No data available for compliance report")
        return None
    
    print("📊 Generating compliance report...")
    
    # Create DataFrame
    df = pd.DataFrame(updated_values, columns=[
        'Khu vực', 'Thiết bị', 'Phương pháp', 'Tần suất (ngày)',
        'Ngày vệ sinh gần nhất', 'Ngày kế hoạch vệ sinh tiếp theo', 'Trạng thái', 'Đang chứa sản phẩm'
    ])
    
    # Calculate compliance metrics
    total_equipment = len(df)
    status_counts = df['Trạng thái'].value_counts()
    
    compliant = status_counts.get('Bình thường', 0) + status_counts.get('Sắp đến hạn', 0)
    compliance_rate = (compliant / total_equipment * 100) if total_equipment > 0 else 0
    
    critical_equipment = len(df[(df['Trạng thái'] == 'Quá hạn') & 
                                (df['Đang chứa sản phẩm'].notna()) & 
                                (df['Đang chứa sản phẩm'].str.strip() != '')])
    
    # Area-wise analysis
    area_analysis = df.groupby('Khu vực')['Trạng thái'].value_counts().unstack(fill_value=0)
    
    # Generate report
    report = {
        'timestamp': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'total_equipment': total_equipment,
        'compliance_rate': round(compliance_rate, 2),
        'status_breakdown': status_counts.to_dict(),
        'critical_equipment': critical_equipment,
        'area_analysis': area_analysis.to_dict() if not area_analysis.empty else {},
        'recommendations': []
    }
    
    # Add recommendations
    if compliance_rate < 80:
        report['recommendations'].append("Compliance rate below 80% - immediate action required")
    
    if critical_equipment > 0:
        report['recommendations'].append(f"🚨 {critical_equipment} critical equipment (overdue + has product) needs immediate attention")
    
    overdue_count = status_counts.get('Quá hạn', 0)
    if overdue_count > 0:
        report['recommendations'].append(f"⚠️ {overdue_count} equipment overdue for cleaning")
    
    print(f"✅ Compliance report generated:")
    print(f"  - Overall compliance rate: {compliance_rate:.1f}%")
    print(f"  - Critical equipment: {critical_equipment}")
    print(f"  - Total overdue: {overdue_count}")
    
    return report

def export_to_csv(sheets_data, export_dir="exports"):
    """
    Export all sheets to CSV files for external analysis
    """
    try:
        # Create export directory if it doesn't exist
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        exported_files = []
        
        for sheet_name, df in sheets_data.items():
            if not df.empty:
                filename = f"{sheet_name.replace(' ', '_')}_{timestamp}.csv"
                filepath = os.path.join(export_dir, filename)
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                exported_files.append(filepath)
                print(f"✅ Exported {sheet_name} to {filepath}")
        
        print(f"📁 Exported {len(exported_files)} files to {export_dir}")
        return exported_files
        
    except Exception as e:
        print(f"❌ Error exporting to CSV: {str(e)}")
        return []

def print_system_info():
    """
    Print system information and requirements
    """
    print("🔧 CIP Cleaning Management System")
    print("=" * 50)
    print(f"📅 Current Date: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"🐍 Python Version: {sys.version}")
    print("\n📋 Required Dependencies:")
    required_packages = [
        'pandas', 'openpyxl', 'requests', 'matplotlib', 
        'msal', 'smtplib', 'email', 'datetime'
    ]
    
    for package in required_packages:
        try:
            __import__(package)
            print(f"  ✅ {package}")
        except ImportError:
            print(f"  ❌ {package} - NOT INSTALLED")
    
    print(f"\n🔑 Environment Variables:")
    env_vars = ['SHAREPOINT_ACCESS_TOKEN', 'SHAREPOINT_REFRESH_TOKEN', 'GITHUB_TOKEN']
    for var in env_vars:
        if os.environ.get(var):
            print(f"  ✅ {var}: {'*' * 20}...{os.environ.get(var)[-5:]}")
        else:
            print(f"  ❌ {var}: NOT SET")
    
    print(f"\n📁 SharePoint Configuration:")
    print(f"  - Tenant ID: {SHAREPOINT_CONFIG['tenant_id']}")
    print(f"  - Site Name: {SHAREPOINT_CONFIG['site_name']}")
    print(f"  - File ID: {CIP_PLAN_FILE_ID}")

# Run the update if executed directly
if __name__ == "__main__":
    print("🚀 Starting CIP Cleaning Management System...")
    print_system_info()
    
    # Ask user what to do
    print("\n🎯 Available Operations:")
    print("1. Full system update (SharePoint sync + email reports)")
    print("2. Test Excel formatting only")
    print("3. Create local backup")
    print("4. Validate data integrity")
    print("5. Export to CSV")
    print("6. Generate compliance report")
    
    try:
        # For automated runs, default to full update
        choice = os.environ.get('RUN_MODE', '1')
        
        if choice == '1':
            print("\n🔄 Running full system update...")
            success = run_update()
            if success:
                print("✅ CIP Cleaning automation completed successfully!")
            else:
                print("❌ CIP Cleaning automation failed!")
                sys.exit(1)
                
        elif choice == '2':
            print("\n🧪 Testing Excel formatting...")
            test_excel_formatting()
            
        elif choice == '3':
            print("\n💾 Creating local backup...")
            # This would need real data - placeholder for now
            print("ℹ️ This option requires real SharePoint data")
            
        elif choice == '4':
            print("\n🔍 Validating data integrity...")
            # This would need real data - placeholder for now
            print("ℹ️ This option requires real SharePoint data")
            
        elif choice == '5':
            print("\n📄 Exporting to CSV...")
            # This would need real data - placeholder for now
            print("ℹ️ This option requires real SharePoint data")
            
        elif choice == '6':
            print("\n📊 Generating compliance report...")
            # This would need real data - placeholder for now
            print("ℹ️ This option requires real SharePoint data")
            
        else:
            print("❌ Invalid choice, defaulting to full update...")
            success = run_update()
            
    except KeyboardInterrupt:
        print("\n🛑 Operation cancelled by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Unexpected error: {str(e)}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        sys.exit(1)
