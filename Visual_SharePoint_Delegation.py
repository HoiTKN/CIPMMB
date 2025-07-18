import msal
import os
import json
import requests
import time
from datetime 
import datetimeimport msal
import os
import json
import requests
import time
from datetime import datetime
from io import BytesIO
import openpyxl

# SharePoint file IDs
SHAREPOINT_FILE_IDS = {
    'sample_id': '8220CAEA-0CD9-585B-D483-DE0A82A98564',
    'data_sx': '6CB4A738-1EDD-4BC4-9996-43A815D3F5CF', 
    'cf_data_output': 'E1B65B6F-6A53-52E0-1BB3-3BCA75A32F63'
}

class MSALSharePointProcessor:
    def __init__(self):
        """Initialize MSAL SharePoint processor"""
        # App configuration
        self.client_id = os.getenv('CLIENT_ID', '076541aa-c734-405e-8518-ed52b67f8cbd')
        self.tenant_id = os.getenv('TENANT_ID', '81060475-7e7f-4ede-8d8d-bf61f53ca528')
        self.client_secret = os.getenv('CLIENT_SECRET')
        self.authority = f"https://login.microsoftonline.com/{self.tenant_id}"
        
        # SharePoint specific scopes
        self.scopes = [
            "https://graph.microsoft.com/Sites.Read.All",
            "https://graph.microsoft.com/Sites.ReadWrite.All",
            "https://graph.microsoft.com/Files.ReadWrite.All"
        ]
        
        self.sharepoint_site_url = os.getenv('SHAREPOINT_SITE_URL', 
                                           'https://masangroup.sharepoint.com/sites/MCH.MMB.QA')
        
        # Token cache setup
        self.cache = msal.SerializableTokenCache()
        self.cache_file = os.getenv('MSAL_CACHE_FILE', 'sharepoint_token_cache.json')
        
        # Load existing cache if available
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'r') as f:
                    self.cache.deserialize(f.read())
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Loaded token cache")
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Failed to load cache: {e}")
        
        # Initialize MSAL app - choose type based on available credentials
        if self.client_secret:
            # Confidential client app (with secret)
            self.app = msal.ConfidentialClientApplication(
                client_id=self.client_id,
                client_credential=self.client_secret,
                authority=self.authority,
                token_cache=self.cache
            )
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔑 Using ConfidentialClientApplication with CLIENT_SECRET")
        else:
            # Public client app (without secret)
            self.app = msal.PublicClientApplication(
                client_id=self.client_id,
                authority=self.authority,
                token_cache=self.cache
            )
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔓 Using PublicClientApplication (no CLIENT_SECRET)")
        
        self.access_token = None
        self.drive_id = None
        
        # Authenticate
        if not self.authenticate():
            raise Exception("Authentication failed during initialization")
            
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ MSAL Processor initialized successfully")

    def authenticate(self):
        """Authenticate using MSAL with prioritized strategies"""
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔐 Authenticating with MSAL...")
        
        # Strategy 1: Client Credentials Flow (highest priority when available)
        if self.client_secret and hasattr(self.app, 'acquire_token_for_client'):
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔑 Attempting Client Credentials authentication...")
            try:
                result = self.app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
                
                if result and "access_token" in result:
                    self.access_token = result["access_token"]
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Client Credentials authentication successful")
                    self.save_cache()
                    return True
                else:
                    error_msg = result.get("error_description", "Unknown error")
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Client Credentials failed: {error_msg}")
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Client Credentials error: {str(e)}")
        
        # Strategy 2: Silent authentication (from cache)
        accounts = self.app.get_accounts()
        if accounts and not self.client_secret:  # Only for public client
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔄 Attempting silent authentication...")
            try:
                result = self.app.acquire_token_silent(self.scopes, account=accounts[0])
                
                if result and "access_token" in result:
                    self.access_token = result["access_token"]
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Silent authentication successful")
                    self.save_cache()
                    return True
                else:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Silent authentication failed")
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Silent authentication error: {str(e)}")
        
        # Strategy 3: Use stored token from GitHub Secrets (fallback for CI)
        if self.is_ci_environment():
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🤖 CI environment - checking stored token...")
            stored_token = os.getenv('SHAREPOINT_ACCESS_TOKEN')
            if stored_token:
                self.access_token = stored_token
                # Test if token works
                if self.test_token():
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Stored token is valid")
                    return True
                else:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Stored token is invalid")
        
        # Strategy 4: Device Code Flow (for local development only)
        if not self.is_ci_environment() and not self.client_secret:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 📱 Starting device code flow...")
            try:
                flow = self.app.initiate_device_flow(scopes=self.scopes)
                if "user_code" not in flow:
                    raise Exception("Device flow failed to initiate")
                
                print(f"\n{'='*60}")
                print(f"🔐 DEVICE CODE AUTHENTICATION REQUIRED")
                print(f"{'='*60}")
                print(f"📲 1. Mở browser và truy cập: {flow['verification_uri']}")
                print(f"🔑 2. Nhập mã này: {flow['user_code']}")
                print(f"⏱️  3. Mã sẽ hết hạn sau: {flow.get('expires_in', 900)} giây")
                print(f"{'='*60}\n")
                
                result = self.app.acquire_token_by_device_flow(flow)
                
                if result and "access_token" in result:
                    self.access_token = result["access_token"]
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Device code authentication successful")
                    self.save_cache()
                    return True
                else:
                    error_msg = result.get("error_description", "Unknown error")
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Device code authentication failed: {error_msg}")
                    
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Device code flow error: {str(e)}")
        
        # All strategies failed
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ All authentication strategies failed")
        return False

    def is_ci_environment(self):
        """Check if running in CI/CD environment"""
        ci_indicators = ['GITHUB_ACTIONS', 'CI', 'BUILD_ID', 'JENKINS_URL']
        return any(os.getenv(indicator) for indicator in ci_indicators)

    def test_token(self):
        """Test if current token is valid"""
        try:
            headers = {'Authorization': f'Bearer {self.access_token}'}
            response = requests.get('https://graph.microsoft.com/v1.0/me', 
                                  headers=headers, timeout=10)
            return response.status_code == 200
        except:
            return False

    def save_cache(self):
        """Save token cache to file"""
        if self.cache.has_state_changed:
            try:
                with open(self.cache_file, 'w') as f:
                    f.write(self.cache.serialize())
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 💾 Token cache saved")
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Failed to save cache: {e}")

    def make_graph_request(self, method, url, **kwargs):
        """Make Microsoft Graph API request with token"""
        headers = kwargs.get('headers', {})
        headers['Authorization'] = f'Bearer {self.access_token}'
        headers['Accept'] = 'application/json'
        kwargs['headers'] = headers
        
        response = requests.request(method, url, timeout=60, **kwargs)
        return response

    def get_sharepoint_drive_id(self):
        """Get SharePoint drive ID"""
        if self.drive_id:
            return self.drive_id
            
        try:
            # Parse SharePoint URL
            url_parts = self.sharepoint_site_url.replace('https://', '').split('/')
            hostname = url_parts[0]
            site_path = '/'.join(url_parts[1:])
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔍 Getting SharePoint site: {hostname}:/{site_path}")
            
            # Get site
            site_url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/{site_path}"
            response = self.make_graph_request('GET', site_url)
            response.raise_for_status()
            
            site_data = response.json()
            site_id = site_data['id']
            
            # Get drives
            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
            response = self.make_graph_request('GET', drives_url)
            response.raise_for_status()
            
            drives_data = response.json()
            if drives_data.get('value'):
                self.drive_id = drives_data['value'][0]['id']
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ SharePoint drive ID: {self.drive_id}")
                return self.drive_id
            else:
                raise Exception("No drives found")
                
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Failed to get drive ID: {e}")
            raise

    def download_file(self, file_id, file_name):
        """Download file from SharePoint"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 📥 Downloading {file_name}...")
            
            drive_id = self.get_sharepoint_drive_id()
            url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content"
            
            response = self.make_graph_request('GET', url)
            response.raise_for_status()
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Downloaded {file_name} ({len(response.content)} bytes)")
            return response.content
            
        except Exception as e:
            error_msg = f"Failed to download {file_name}: {str(e)}"
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ {error_msg}")
            raise Exception(error_msg)

    def upload_file(self, file_id, file_content, file_name):
        """Upload file to SharePoint"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 📤 Uploading {file_name}...")
            
            drive_id = self.get_sharepoint_drive_id()
            url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content"
            
            headers = {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = self.make_graph_request('PUT', url, headers=headers, data=file_content)
            response.raise_for_status()
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Uploaded {file_name}")
            return True
            
        except Exception as e:
            error_msg = f"Failed to upload {file_name}: {str(e)}"
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ {error_msg}")
            raise Exception(error_msg)

    def process_files(self):
        """Main processing workflow"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🚀 Starting file processing...")
            
            # Download source files
            sample_id_content = self.download_file(
                SHAREPOINT_FILE_IDS['sample_id'], 
                "Sample ID.xlsx"
            )
            
            data_sx_content = self.download_file(
                SHAREPOINT_FILE_IDS['data_sx'], 
                "Data SX.xlsx"
            )
            
            # Process data
            processed_content = self.process_data(sample_id_content, data_sx_content)
            
            # Upload result
            self.upload_file(
                SHAREPOINT_FILE_IDS['cf_data_output'],
                processed_content,
                "CF data.xlsx"
            )
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Processing completed successfully")
            return True
            
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Processing failed: {str(e)}")
            raise

    def process_data(self, sample_id_content, data_sx_content):
        """Process the Excel data"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔄 Processing data...")
            
            # Load workbooks
            sample_id_wb = openpyxl.load_workbook(BytesIO(sample_id_content))
            data_sx_wb = openpyxl.load_workbook(BytesIO(data_sx_content))
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Loaded Excel files")
            print(f"Sample ID sheets: {sample_id_wb.sheetnames}")
            print(f"Data SX sheets: {data_sx_wb.sheetnames}")
            
            # Create output workbook
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = "CF Data"
            
            # Add processed data
            output_ws['A1'] = "Processing Date"
            output_ws['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            output_ws['A2'] = "Status"
            output_ws['B2'] = "Data processed successfully with MSAL Client Credentials"
            output_ws['A3'] = "Source Files"
            output_ws['B3'] = "Sample ID.xlsx + Data SX.xlsx"
            output_ws['A4'] = "Authentication Method"
            output_ws['B4'] = "MSAL with CLIENT_SECRET"
            
            # TODO: Add your actual data processing logic here
            
            # Save to buffer
            output_buffer = BytesIO()
            output_wb.save(output_buffer)
            output_buffer.seek(0)
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Data processing completed")
            return output_buffer.getvalue()
            
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Data processing error: {str(e)}")
            raise

def main():
    """Main execution function"""
    try:
        print("=" * 60)
        print("🏭 MSAL SHAREPOINT QA DATA PROCESSING")
        print("=" * 60)
        
        # Environment check
        print("🔧 Environment Check:")
        required_vars = ['CLIENT_ID', 'TENANT_ID']
        auth_vars = ['CLIENT_SECRET', 'SHAREPOINT_ACCESS_TOKEN']
        optional_vars = ['SHAREPOINT_SITE_URL']
        
        for var in required_vars:
            status = "✅" if os.getenv(var) else "❌"
            print(f"{status} {var}: {'Found' if os.getenv(var) else 'Missing'}")
        
        for var in auth_vars:
            status = "✅" if os.getenv(var) else "⚠️"
            print(f"{status} {var}: {'Found' if os.getenv(var) else 'Not found'}")
        
        for var in optional_vars:
            status = "✅" if os.getenv(var) else "📋"
            print(f"{status} {var}: {'Found' if os.getenv(var) else 'Using default'}")
        
        print(f"📋 SharePoint File IDs: {SHAREPOINT_FILE_IDS}")
        
        # Check if we have proper authentication
        has_client_secret = bool(os.getenv('CLIENT_SECRET'))
        has_stored_token = bool(os.getenv('SHAREPOINT_ACCESS_TOKEN'))
        
        if has_client_secret:
            print("🔑 Using CLIENT_SECRET for authentication (recommended)")
        elif has_stored_token:
            print("⚠️ Using stored token (may expire)")
        else:
            print("❌ No authentication method available")
            return 1
        
        print("🚀 Initializing processor...")
        processor = MSALSharePointProcessor()
        
        print("📊 Processing files...")
        processor.process_files()
        
        print("✅ All operations completed successfully!")
        return 0
        
    except Exception as e:
        print(f"❌ Critical error: {str(e)}")
        import traceback
        print("Traceback:", traceback.format_exc())
        return 1

if __name__ == "__main__":
    exit_code = main()
    exit(exit_code)
from io import BytesIO
import openpyxl

# SharePoint file IDs
SHAREPOINT_FILE_IDS = {
    'sample_id': '8220CAEA-0CD9-585B-D483-DE0A82A98564',
    'data_sx': '6CB4A738-1EDD-4BC4-9996-43A815D3F5CF', 
    'cf_data_output': 'E1B65B6F-6A53-52E0-1BB3-3BCA75A32F63'
}

class MSALSharePointProcessor:
    def __init__(self):
        """Initialize MSAL SharePoint processor"""
        # App configuration
        self.client_id = os.getenv('CLIENT_ID', '076541aa-c734-405e-8518-ed52b67f8cbd')
        self.tenant_id = os.getenv('TENANT_ID', '81060475-7e7f-4ede-8d8d-bf61f53ca528')
        self.authority = f"https://login.microsoftonline.com/{self.tenant_id}"
        
        # SharePoint specific scopes
        self.scopes = [
            "https://graph.microsoft.com/Sites.Read.All",
            "https://graph.microsoft.com/Sites.ReadWrite.All",
            "https://graph.microsoft.com/Files.ReadWrite.All"
        ]
        
        self.sharepoint_site_url = os.getenv('SHAREPOINT_SITE_URL', 
                                           'https://masangroup.sharepoint.com/sites/MCH.MMB.QA')
        
        # Token cache setup
        self.cache = msal.SerializableTokenCache()
        self.cache_file = os.getenv('MSAL_CACHE_FILE', 'sharepoint_token_cache.json')
        
        # Load existing cache if available
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'r') as f:
                    self.cache.deserialize(f.read())
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Loaded token cache")
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Failed to load cache: {e}")
        
        # Initialize MSAL app
        self.app = msal.PublicClientApplication(
            client_id=self.client_id,
            authority=self.authority,
            token_cache=self.cache
        )
        
        self.access_token = None
        self.drive_id = None
        
        # Authenticate
        if not self.authenticate():
            raise Exception("Authentication failed during initialization")
            
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ MSAL Processor initialized successfully")

    def authenticate(self):
        """Authenticate using MSAL with fallback strategies"""
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔐 Authenticating with MSAL...")
        
        # Strategy 1: Try silent authentication (from cache)
        accounts = self.app.get_accounts()
        if accounts:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔄 Attempting silent authentication...")
            result = self.app.acquire_token_silent(self.scopes, account=accounts[0])
            
            if result and "access_token" in result:
                self.access_token = result["access_token"]
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Silent authentication successful")
                self.save_cache()
                return True
            else:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Silent authentication failed")
        
        # Strategy 2: Device Code Flow (for automation/CI environments)
        if self.is_ci_environment():
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🤖 CI environment detected, using stored token...")
            # In CI, we'll use pre-generated token from GitHub Secrets
            stored_token = os.getenv('SHAREPOINT_ACCESS_TOKEN')
            if stored_token:
                self.access_token = stored_token
                # Test if token works
                if self.test_token():
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Stored token is valid")
                    return True
                else:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Stored token is invalid")
                    return False
            else:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ No stored token found in CI environment")
                return False
        
        # Strategy 3: Interactive Device Code Flow (for local development)
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 📱 Starting device code flow...")
        try:
            flow = self.app.initiate_device_flow(scopes=self.scopes)
            if "user_code" not in flow:
                raise Exception("Device flow failed to initiate")
            
            print(f"\n{'='*60}")
            print(f"🔐 DEVICE CODE AUTHENTICATION REQUIRED")
            print(f"{'='*60}")
            print(f"📲 1. Mở browser và truy cập: {flow['verification_uri']}")
            print(f"🔑 2. Nhập mã này: {flow['user_code']}")
            print(f"⏱️  3. Mã sẽ hết hạn sau: {flow.get('expires_in', 900)} giây")
            print(f"{'='*60}\n")
            
            result = self.app.acquire_token_by_device_flow(flow)
            
            if result and "access_token" in result:
                self.access_token = result["access_token"]
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Device code authentication successful")
                self.save_cache()
                return True
            else:
                error_msg = result.get("error_description", "Unknown error")
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Device code authentication failed: {error_msg}")
                return False
                
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Device code flow error: {str(e)}")
            return False

    def is_ci_environment(self):
        """Check if running in CI/CD environment"""
        ci_indicators = ['GITHUB_ACTIONS', 'CI', 'BUILD_ID', 'JENKINS_URL']
        return any(os.getenv(indicator) for indicator in ci_indicators)

    def test_token(self):
        """Test if current token is valid"""
        try:
            headers = {'Authorization': f'Bearer {self.access_token}'}
            response = requests.get('https://graph.microsoft.com/v1.0/me', 
                                  headers=headers, timeout=10)
            return response.status_code == 200
        except:
            return False

    def save_cache(self):
        """Save token cache to file"""
        if self.cache.has_state_changed:
            try:
                with open(self.cache_file, 'w') as f:
                    f.write(self.cache.serialize())
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 💾 Token cache saved")
            except Exception as e:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠️ Failed to save cache: {e}")

    def make_graph_request(self, method, url, **kwargs):
        """Make Microsoft Graph API request with token"""
        headers = kwargs.get('headers', {})
        headers['Authorization'] = f'Bearer {self.access_token}'
        headers['Accept'] = 'application/json'
        kwargs['headers'] = headers
        
        response = requests.request(method, url, timeout=60, **kwargs)
        return response

    def get_sharepoint_drive_id(self):
        """Get SharePoint drive ID"""
        if self.drive_id:
            return self.drive_id
            
        try:
            # Parse SharePoint URL
            url_parts = self.sharepoint_site_url.replace('https://', '').split('/')
            hostname = url_parts[0]
            site_path = '/'.join(url_parts[1:])
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔍 Getting SharePoint site: {hostname}:/{site_path}")
            
            # Get site
            site_url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/{site_path}"
            response = self.make_graph_request('GET', site_url)
            response.raise_for_status()
            
            site_data = response.json()
            site_id = site_data['id']
            
            # Get drives
            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
            response = self.make_graph_request('GET', drives_url)
            response.raise_for_status()
            
            drives_data = response.json()
            if drives_data.get('value'):
                self.drive_id = drives_data['value'][0]['id']
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ SharePoint drive ID: {self.drive_id}")
                return self.drive_id
            else:
                raise Exception("No drives found")
                
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Failed to get drive ID: {e}")
            raise

    def download_file(self, file_id, file_name):
        """Download file from SharePoint"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 📥 Downloading {file_name}...")
            
            drive_id = self.get_sharepoint_drive_id()
            url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content"
            
            response = self.make_graph_request('GET', url)
            response.raise_for_status()
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Downloaded {file_name} ({len(response.content)} bytes)")
            return response.content
            
        except Exception as e:
            error_msg = f"Failed to download {file_name}: {str(e)}"
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ {error_msg}")
            raise Exception(error_msg)

    def upload_file(self, file_id, file_content, file_name):
        """Upload file to SharePoint"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 📤 Uploading {file_name}...")
            
            drive_id = self.get_sharepoint_drive_id()
            url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content"
            
            headers = {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = self.make_graph_request('PUT', url, headers=headers, data=file_content)
            response.raise_for_status()
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Uploaded {file_name}")
            return True
            
        except Exception as e:
            error_msg = f"Failed to upload {file_name}: {str(e)}"
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ {error_msg}")
            raise Exception(error_msg)

    def process_files(self):
        """Main processing workflow"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🚀 Starting file processing...")
            
            # Download source files
            sample_id_content = self.download_file(
                SHAREPOINT_FILE_IDS['sample_id'], 
                "Sample ID.xlsx"
            )
            
            data_sx_content = self.download_file(
                SHAREPOINT_FILE_IDS['data_sx'], 
                "Data SX.xlsx"
            )
            
            # Process data
            processed_content = self.process_data(sample_id_content, data_sx_content)
            
            # Upload result
            self.upload_file(
                SHAREPOINT_FILE_IDS['cf_data_output'],
                processed_content,
                "CF data.xlsx"
            )
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Processing completed successfully")
            return True
            
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Processing failed: {str(e)}")
            raise

    def process_data(self, sample_id_content, data_sx_content):
        """Process the Excel data"""
        try:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 🔄 Processing data...")
            
            # Load workbooks
            sample_id_wb = openpyxl.load_workbook(BytesIO(sample_id_content))
            data_sx_wb = openpyxl.load_workbook(BytesIO(data_sx_content))
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Loaded Excel files")
            print(f"Sample ID sheets: {sample_id_wb.sheetnames}")
            print(f"Data SX sheets: {data_sx_wb.sheetnames}")
            
            # Create output workbook
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = "CF Data"
            
            # Add processed data
            output_ws['A1'] = "Processing Date"
            output_ws['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            output_ws['A2'] = "Status"
            output_ws['B2'] = "Data processed successfully with MSAL authentication"
            output_ws['A3'] = "Source Files"
            output_ws['B3'] = "Sample ID.xlsx + Data SX.xlsx"
            
            # TODO: Add your actual data processing logic here
            
            # Save to buffer
            output_buffer = BytesIO()
            output_wb.save(output_buffer)
            output_buffer.seek(0)
            
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✅ Data processing completed")
            return output_buffer.getvalue()
            
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ❌ Data processing error: {str(e)}")
            raise

def generate_token_for_github_secrets():
    """Helper function to generate token for GitHub Secrets (run locally)"""
    print("🔧 GITHUB SECRETS TOKEN GENERATOR")
    print("=" * 50)
    
    processor = MSALSharePointProcessor()
    
    if processor.access_token:
        print(f"\n✅ SUCCESS! Copy this token to GitHub Secrets:")
        print(f"SHAREPOINT_ACCESS_TOKEN={processor.access_token}")
        print(f"\n📋 Steps:")
        print(f"1. Go to GitHub repo → Settings → Secrets and variables → Actions")
        print(f"2. Edit SHAREPOINT_ACCESS_TOKEN")
        print(f"3. Paste the token above")
        return processor.access_token
    else:
        print("❌ Failed to generate token")
        return None

def main():
    """Main execution function"""
    try:
        print("=" * 60)
        print("🏭 MSAL SHAREPOINT QA DATA PROCESSING")
        print("=" * 60)
        
        # Check if we need to generate token for GitHub
        if len(os.sys.argv) > 1 and os.sys.argv[1] == "--generate-token":
            return generate_token_for_github_secrets()
        
        # Normal processing
        processor = MSALSharePointProcessor()
        processor.process_files()
        
        print("✅ All operations completed successfully!")
        return 0
        
    except Exception as e:
        print(f"❌ Critical error: {str(e)}")
        import traceback
        print("Traceback:", traceback.format_exc())
        return 1

if __name__ == "__main__":
    exit_code = main()
    if isinstance(exit_code, str):  # Token generation case
        print(f"\nGenerated token: {exit_code[:50]}...")
    else:
        exit(exit_code)
